import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(test_results, output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()
    
    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Styling
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=12, bold=True, color="000000")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)
    
    fill_blue_title = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_pass = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_fail = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Title Block
    ws_summary.merge_cells("A1:D2")
    ws_summary["A1"] = "Road Rescue System — Test Automation Report"
    ws_summary["A1"].font = font_title
    ws_summary["A1"].fill = fill_blue_title
    ws_summary["A1"].alignment = align_center
    
    # Execution Info
    ws_summary["A4"] = "Execution Overview"
    ws_summary["A4"].font = font_section
    
    headers = ["Metric", "Value"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=5, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_blue_title
        cell.alignment = align_center
        cell.border = border_thin
        
    total_tests = len(test_results)
    passed_tests = sum(1 for r in test_results if r["status"] == "PASSED")
    failed_tests = total_tests - passed_tests
    pass_pct = f"{(passed_tests / total_tests * 100):.1f}%" if total_tests > 0 else "0.0%"
    
    metrics = [
        ("Total Tests Run", total_tests),
        ("Passed Tests", passed_tests),
        ("Failed Tests", failed_tests),
        ("Pass Percentage", pass_pct)
    ]
    
    for row_idx, (metric, val) in enumerate(metrics, 6):
        c1 = ws_summary.cell(row=row_idx, column=1, value=metric)
        c2 = ws_summary.cell(row=row_idx, column=2, value=val)
        c1.font = font_data
        c2.font = font_bold
        c1.border = border_thin
        c2.border = border_thin
        c2.alignment = align_center
        if metric == "Passed Tests" and passed_tests > 0:
            c2.fill = fill_pass
        elif metric == "Failed Tests" and failed_tests > 0:
            c2.fill = fill_fail

    # Auto-adjust column width
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 15)

    # 2. Detailed Test Results Sheet
    ws_details = wb.create_sheet(title="Test Details")
    ws_details.views.sheetView[0].showGridLines = True
    
    detail_headers = ["Test ID", "Test Suite", "Test Case Name", "Status", "Duration (s)", "Error Message"]
    for col_idx, h in enumerate(detail_headers, 1):
        cell = ws_details.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_blue_title
        cell.alignment = align_center
        cell.border = border_thin
        
    for row_idx, r in enumerate(test_results, 2):
        row_data = [
            f"TC_{row_idx-1:03d}",
            r["suite"],
            r["name"],
            r["status"],
            round(r["duration"], 2),
            r.get("error", "")
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_details.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_thin
            
            if col_idx == 4: # Status Column
                cell.alignment = align_center
                if val == "PASSED":
                    cell.fill = fill_pass
                    cell.font = font_bold
                else:
                    cell.fill = fill_fail
                    cell.font = font_bold
            elif col_idx == 5: # Duration
                cell.alignment = align_center
                
    for col in ws_details.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_details.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(output_path)
    print(f"[REPORTS] Excel Automation Report written: {output_path}")
